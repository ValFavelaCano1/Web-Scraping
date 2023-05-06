from urllib.request import urlopen, Request 
from bs4 import BeautifulSoup

import openpyxl as xl
from openpyxl.styles import Font, Alignment, numbers

import keys
from twilio.rest import Client

client = Client(keys.accountSID, keys.authToken)

TwilioNumber = '+19703605354'
mycellphone = '+19703903557'

url = 'https://www.investing.com/crypto/currencies'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title.text)


#Start Excel Workbook
currencies_rows = soup.findAll("tr")

wb = xl.Workbook()
ws = wb.active

ws.title = 'Top 5 CryptoCurrencies'

ws['A1'] = 'Rank'
ws['B1'] = 'Name'
ws['C1'] = 'Symbol'
ws['D1'] = 'Current Price'
ws['E1'] = '24h% Change'
ws['F1'] = 'Price Change'

#Start Webscraping to Sheet
for row in range(1,6):
    td = currencies_rows[row].findAll("td")
    if td:
        rank = td[0].text
        name = td[2].text
        symbol = td[3].text
        current_price = round(float(td[4].text.replace(",","").replace("$","")), 4)
        percent_change = round(float(td[8].text.replace("%","").replace("+","")), 4)

        price_change = round((current_price / 100) * percent_change, 2)
        
        ws['A' + str(row + 1)] = rank
        ws['B' + str(row + 1)] = name
        ws['C' + str(row + 1)] = symbol
        ws['D' + str(row + 1)] = '$' + str(current_price)
        ws['E' + str(row + 1)] = str(percent_change) + '%'
        ws['F' + str(row + 1)] = '$' + str(price_change)


        #Alert Message
        if'BTC' in symbol or 'ETH' in symbol:
            if price_change > 0 and price_change <= 5:
                message = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                     body= "The price of" + name.upper() + "has increased by " + "$" + str(price_change))
            if price_change < 0 and abs(price_change) <= 5:
                message = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                   body= "The price of" + name.upper() + "has decreased by " + "$" + str(price_change))

ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 21
ws.column_dimensions['E'].width = 21
ws.column_dimensions['F'].width = 20

header_font = Font(size=20, bold= True, color='40957A')
header_alignment = Alignment(horizontal= 'center')
cell_alignment = Alignment(horizontal='right')
cell_font = Font(size=16)

'''
for cell in ws['D']:
    cell.alignment = cell_alignment

for cell in ws['E']:
    cell.alignment = cell_alignment

for cell in ws['F']:
    cell.alignment = cell_alignment
'''

for cell in ws[1:1]:
   cell.font = header_font
   cell.alignment = header_alignment

for row in ws['A2:F6']:
    for cell in row:
        cell.font = cell_font

for row in ws['D2:F6']:
    for cell in row:
        cell.alignment = cell_alignment

for cell in ws['A']:
    cell.alignment = Alignment(horizontal='center')

for cell in ws['D']:
    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

for cell in ws['F']:
    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

#Save Excel Workbook
wb.save("Top 5 CryptoCurrencies.xlsx")

